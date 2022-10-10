# 读取excel数据 xlrd xlwt xlutils pandas openpyxl

# openpyxl操作excel
# 1.安装 pip install openpyxl  pip list
# 操作excel:
# 打开excel
# 读取excel数据

import openpyxl

# 打开excel
workbook=openpyxl.load_workbook('./data/bb.xlsx')
# print(workbook)
# 拿表单 指定表单拿里面的数据
# sheet=workbook['Sheet3']
# print(sheet)
# 读取excel里面的数据 for
# for i in sheet.values:
#     print(i)

# workbook['Sheet3'].values

# 拿到所有的表单数据
# sheet=workbook.sheetnames
# # print(sheet)
# # 拿到所有的表单数据
# # 表单先都拿出来  循环列表
# # 表单里面的值  workbook[Sheet1].values  workbook[Sheet2].values  Sheet3
# for i in sheet:
#     print(i)
#     for j in workbook[i].values:
#         print(j)

# 不要所有的数据  只要某一个值
# 指定表单
sheet=workbook['Sheet3']
# 指定单元格的数据
# id=sheet['A1'].value
# print(id)
# # 指定单元格的数据 行和列
# id=sheet.cell(row=1,column=1).value
# print(id)

# 写入数据
# sheet['E1'].value='预期结果3'
# workbook.save('./data/bb.xlsx')