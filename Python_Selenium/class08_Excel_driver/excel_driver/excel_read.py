'''
    excel文件读取类。用于实现测试用例文件的读取与执行。
'''
import openpyxl

# 解析测试用例中测试参数单元格的内容，并转换为字典的形态返回
from class07_web_keys.web_keys import Keys


def arguments(value):
    data = dict()
    # 如果value有值，进行切分
    if value:
        str_temp = value.split(';')
        for temp in str_temp:
            t = temp.split('=', 1)
            data[t[0]] = t[1]
    # 如果value没有值，就不做任何操作。
    else:
        pass
    return data


# 获取excel中的内容
excel = openpyxl.load_workbook('../data/自动化测试用例demo.xlsx')

# sheet = excel['Sheet1']
# 获取所有的sheet页，来执行里面的测试内容
for name in excel.sheetnames:
    sheet = excel[name]
    print('**********正在执行{}Sheet页*********'.format(name))
    for values in sheet.values:
        # 获取测试用例的正文内容
        if type(values[0]) is int:
            # 用例描述可以用于日志的输出
            print('*****************正在执行：{}*****************'.format(values[3]))
            # print(values)
            # 参数的处理:通过一个dict来接收所有的参数内容。便于定值不定长的传参形态
            # 参数最终形态：'type_=Chrome' 改变为 {type_: 'Chrome'}
            data = arguments(values[2])
            # print(data)
            '''
                调用的函数是values[1]，这是固定的。
                    操作行为的调用分为以下几种不同类型：
                        1. 实例化
                        2. 基于实例化对象进行的操作行为
                        3. 断言机制：有预期与实际的对比，以及有单元格测试结果的写入
            '''
            '''
                第一行open_browser
                第二行open:getattr(key,'open')(**data)
                    key.open(**data)
                    key.open(url='http://www.baidu.com')
                第三行input:getattr(key,'input')(**data)
                    key.input(**data)
                    key.input(by='id',value='kw',txt='虚竹的Excel')
                第四行click
            '''
            # 实例化操作
            if values[1] == 'open_browser':
                key = Keys(**data)
            # 断言行为:基于断言的返回结果来判定测试的成功失败，并进行写入操作
            elif 'assert' in values[1]:
                status = getattr(key, values[1])(expected=values[4], **data)
                #
                if status:
                    sheet.cell(row=values[0] + 2, column=6).value = 'Pass'
                else:
                    sheet.cell(row=values[0] + 2, column=6).value = 'Failed'
                # 保存Excel：放在这里以确保每一次写入都可以被保存，避免因为代码报错而未保存之前的测试结果
                excel.save('../data/自动化测试用例demo.xlsx')
            # 常规操作行为
            else:
                getattr(key, values[1])(**data)
